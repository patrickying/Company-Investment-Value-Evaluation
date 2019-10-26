from openpyxl import load_workbook
import pandas as pd
import numpy as np
import math

DIR_PATH = '../data/'  # training data is in 'data/'

# change the number to column in excel
def num_to_column(num, word=""):
    if num <= 26:
        return word + chr(num + 64)
    else:
        temp = num
        count = 0
        while temp > 26:
            temp = (temp-1) // 26
            count += 1

        if num % 26**count == 0:
            num = 26
        else:
            num = num % 26**count
        result = num_to_column(num, word + chr(temp + 64))
        return result


def is_number(s):
    try:
        float(s)
        return True
    except Exception:
        pass
    return False

# write to excel
def WriteCsv(excel_name, target, data, stat):
    wbWrite = load_workbook(DIR_PATH + 'Input.xlsx')
    wbWritesheet = wbWrite.get_sheet_by_name('Sheet1')
    wb = load_workbook(DIR_PATH + excel_name)
    sheet = wb.get_sheet_by_name('Sheet')
    label = load_workbook(DIR_PATH + '企业评分.xlsx')
    labelsheet = label.get_sheet_by_name('Sheet')

    companyNum = set()
    for row in labelsheet.rows:
        if is_number(row[0].value):
            companyNum.add(int(row[0].value))

    line = 2
    max_column = wbWritesheet.max_column
    titleWrote = False
    miss_count = [0]*len(target)
    for rowIndex, row in enumerate(sheet.rows):
        if not titleWrote:
            titleWrote = True
            for Position, num in enumerate(target):
                wbWritesheet[num_to_column(max_column + Position + 1) + '1'] = row[num].value + stat

        if is_number(row[0].value):  # 如果企业编号是数字
            if int(row[0].value) in companyNum:
                companyNum.remove(int(row[0].value))
                for Position, num in enumerate(target):
                    if is_number(row[num].value):
                        wbWritesheet[num_to_column(max_column + Position + 1) + str(line)] = row[num].value
                    else:
                        miss_count[Position] += 1
                        wbWritesheet[num_to_column(max_column + Position + 1) + str(line)] = data[Position]
                line += 1
    # print(miss_count)
    wbWrite.save(DIR_PATH + 'Input.xlsx')


# fill the missing value with average
def AddAverage(excel_name, target):
    wb = load_workbook(DIR_PATH + excel_name)
    sheet = wb.get_sheet_by_name('Sheet')

    sum = [0.0] * len(target)
    count = [0.0] * len(target)

    temp = "企业编号"
    for row in sheet.rows:
        if temp != row[0].value:
            temp = row[0].value
            for index, cell in enumerate(row):
                if index in target and is_number(cell.value):
                    Position = target.index(index)
                    sum[Position] += float(row[index].value)
                    count[Position] += 1

    for x in range(len(count)):
        if count[x] == 0:
            count[x] = 1

    avg = [sum[x]/count[x] for x in range(len(target))]
    WriteCsv(excel_name, target, avg, "_Avg")


# fill the missing value with median
def AddMedian(excel_name, target):
    wb = load_workbook(DIR_PATH + excel_name)
    sheet = wb.get_sheet_by_name('Sheet')

    mediam = []
    for x in range(len(target)):
        mediam.append([])

    temp = "企业编号"
    for row in sheet.rows:
        if temp != row[0].value:
            temp = row[0].value
            for index, cell in enumerate(row):
                if index in target and is_number(cell.value):
                    Position = target.index(index)
                    mediam[Position].append(float(row[index].value))

    med = []
    for x in range(len(mediam)):
        temp_med = sorted(mediam[x])
        if len(temp_med) >= 1:
            med.append(temp_med[(len(temp_med)-1)//2])
        else:
            med.append(0)

    WriteCsv(excel_name, target, med, "_Med")


# change category feature to onehot
def OneHot(excel_name, targetList):
    wbWrite = load_workbook(DIR_PATH + 'Input.xlsx')
    wbWritesheet = wbWrite.get_sheet_by_name('Sheet1')
    wb = load_workbook(DIR_PATH + excel_name)
    sheet = wb.get_sheet_by_name('Sheet')
    label = load_workbook(DIR_PATH + '企业评分.xlsx')
    labelsheet = label.get_sheet_by_name('Sheet')

    for target in targetList:
        companyNum = []
        companyOneHot = {}
        for row in labelsheet.rows:
            if is_number(row[0].value):
                companyOneHot[int(row[0].value)] = [0]
                companyNum.append(int(row[0].value))

        category = set()
        temp = "企业编号"
        for row in sheet.rows:
            if temp != row[0].value:
                temp = row[0].value
                if row[target].value:
                    category.add(row[target].value)
                else:
                    category.add('None')
        print(category)
        category = sorted(list(category))
        line = 2
        max_column = wbWritesheet.max_column

        titleWrote = False
        for rowIndex, row in enumerate(sheet.rows):
            if not titleWrote:
                titleWrote = True
                for Position, num in enumerate(category):
                    wbWritesheet[num_to_column(max_column + Position + 1) + '1'] = row[target].value+"_"+num + "OneHot"

            onehot = [0] * len(category)
            if is_number(row[0].value):  # 如果企业编号是数字
                if int(row[0].value) in companyNum and len(companyOneHot[int(row[0].value)]) == 1:  # 如果企业编号符合目前公司代码

                    if row[target].value:
                        onehot[category.index(row[target].value)] = 1
                    else:
                        onehot[category.index('None')] = 1
                    companyOneHot[int(row[0].value)] = onehot

        onehot = [0] * len(category)
        if 'None' in category:
            onehot[category.index('None')] = 1

        for x in companyNum:
            if len(companyOneHot[x]) == 1:
                for Position, num in enumerate(onehot):
                    wbWritesheet[num_to_column(max_column + Position + 1) + str(line)] = num
            else:
                for Position, num in enumerate(companyOneHot[x]):
                    wbWritesheet[num_to_column(max_column + Position + 1) + str(line)] = num
            line += 1

        wbWrite.save(DIR_PATH + 'Input.xlsx')


# the company is existed
def WriteExist(excel_name):
    wbWrite = load_workbook(DIR_PATH + 'Input.xlsx')
    wbWritesheet = wbWrite.get_sheet_by_name('Sheet1')
    wb = load_workbook(DIR_PATH+ excel_name)
    sheet = wb.get_sheet_by_name('Sheet')
    label = load_workbook(DIR_PATH + '企业评分.xlsx')
    labelsheet = label.get_sheet_by_name('Sheet')

    companyNum = []
    for row in labelsheet.rows:
        if is_number(row[0].value):
            companyNum.append(int(row[0].value))

    line = 2
    max_column = wbWritesheet.max_column
    titleWrote = False

    targetNum = set()
    for rowIndex, row in enumerate(sheet.rows):
        if not titleWrote:
            titleWrote = True
            wbWritesheet[num_to_column(max_column + 1) + '1'] = excel_name

        if is_number(row[0].value):  # 如果企业编号是数字
            targetNum.add(int(row[0].value))

    for x in companyNum:
        if x in targetNum:
            wbWritesheet[num_to_column(max_column + 1) + str(line)] = 1
        else:
            wbWritesheet[num_to_column(max_column + 1) + str(line)] = 0
        line += 1

    wbWrite.save(DIR_PATH + 'Input.xlsx')


# count the company occurrences
def WriteCount(excel_name, filter_index=-1, filter_limit=0, islog=False):
    wbWrite = load_workbook(DIR_PATH + 'Input.xlsx')
    wbWritesheet = wbWrite.get_sheet_by_name('Sheet1')
    wb = load_workbook(DIR_PATH + excel_name)
    sheet = wb.get_sheet_by_name('Sheet')
    label = load_workbook(DIR_PATH + '企业评分.xlsx')
    labelsheet = label.get_sheet_by_name('Sheet')

    companyNum = []
    companyCount = {}
    for row in labelsheet.rows:
        if is_number(row[0].value):
            if int(row[0].value) not in companyNum:
                companyNum.append(int(row[0].value))
                companyCount[int(row[0].value)] = 0

    line = 2
    max_column = wbWritesheet.max_column
    titleWrote = False

    for rowIndex, row in enumerate(sheet.rows):
        if not titleWrote:
            titleWrote = True
            if filter_index == -1:
                wbWritesheet[num_to_column(max_column + 1) + '1'] = excel_name + '_Count'
            else:
                wbWritesheet[num_to_column(max_column + 1) + '1'] = excel_name + '_Count' + str(filter_limit)

        if is_number(row[0].value) and int(row[0].value) in companyNum:  # 如果企业编号是数字
            if filter_index != -1:
                if row[filter_index].value is not None and row[filter_index].value >= filter_limit:
                    companyCount[int(row[0].value)] += 1
            else:
                companyCount[int(row[0].value)] += 1

    for x in companyNum:
        if islog:
            wbWritesheet[num_to_column(max_column + 1) + str(line)] = math.log(companyCount[x]+1, 10)
        else:
            wbWritesheet[num_to_column(max_column + 1) + str(line)] = companyCount[x]
        line += 1

    wbWrite.save(DIR_PATH + 'Input.xlsx')


# change season data to year data by average
def AddAverageYear(excel_name, targetList):
    wbWrite = load_workbook(DIR_PATH + 'Input.xlsx')
    wbWritesheet = wbWrite.get_sheet_by_name('Sheet1')
    wb = load_workbook(DIR_PATH + excel_name)
    sheet = wb.get_sheet_by_name('Sheet')
    label = load_workbook(DIR_PATH + '企业评分.xlsx')
    labelsheet = label.get_sheet_by_name('Sheet')

    for target in targetList:
        companyNum = []
        companyCount = {}
        for row in labelsheet.rows:
            if is_number(row[0].value):
                if int(row[0].value) not in companyNum:
                    companyNum.append(int(row[0].value))
                    companyCount[int(row[0].value)] = [0.0, 0.0]

        sum = 0.0
        count = 0.0
        max_column = wbWritesheet.max_column

        temp = "企业编号"
        titleWrote = False
        for row in sheet.rows:
            if not titleWrote:
                titleWrote = True
                wbWritesheet[num_to_column(max_column + 1) + '1'] = row[target].value + "_Year"
            if is_number(row[0].value) and int(row[0].value) in companyNum and companyCount[int(int(row[0].value))][0] < 4:
                if is_number(row[target].value):
                    companyCount[int(int(row[0].value))][0] += 1
                    companyCount[int(int(row[0].value))][1] += float(row[target].value)

        for x in companyNum:
            if companyCount[x][0] != 0:
                companyCount[x][1] /= companyCount[x][0]
                sum += companyCount[x][1]
                count += 1

        if count == 0:
            avg = 0
        else:
            avg = sum / count

        line = 2
        for x in companyNum:
            if companyCount[x][0] != 0:
                wbWritesheet[num_to_column(max_column + 1) + str(line)] = companyCount[x][1]
            else:
                wbWritesheet[num_to_column(max_column + 1) + str(line)] = avg
            line += 1

        wbWrite.save(DIR_PATH + 'Input.xlsx')


def xlsx_to_csv_pd():
    data_xls = pd.read_excel(DIR_PATH +'Input.xlsx', index_col=0)
    data_xls.to_csv(DIR_PATH + 'Input.csv', encoding='utf-8_sig')


if __name__ == "__main__":
    # AddAverage('上市公司财务信息-每股指标.xlsx', [x for x in range(3, 10)])
    # AddAverage('上市信息财务信息-财务风险指标.xlsx', [x for x in range(3, 7)])
    # AddAverage('上市信息财务信息-成长能力指标.xlsx', [x for x in range(13, 23)])
    # AddAverage('上市信息财务信息-现金流量表.xlsx', [x for x in range(3, 27)])
    # AddAverage('上市信息财务信息资产负债表.xlsx', [x for x in range(3, 23)])
    # AddAverage('上市信息财务信息运营能力指标.xlsx', [x for x in range(3, 6)])
    # AddAverage('上市信息财务信息盈利能力指标.xlsx', [x for x in range(9, 15)])
    # AddAverage('工商基本信息表.xlsx', [x for x in range(1,2)])
    #
    # print("AddMedium")
    # AddMedian('上市公司财务信息-每股指标.xlsx', [x for x in range(3, 10)])
    # AddMedian('上市信息财务信息-财务风险指标.xlsx', [x for x in range(3, 7)])
    # AddMedian('上市信息财务信息-成长能力指标.xlsx', [x for x in range(13, 23)])
    # AddMedian('上市信息财务信息-现金流量表.xlsx', [x for x in range(3, 27)])
    # AddMedian('上市信息财务信息资产负债表.xlsx', [x for x in range(3, 23)])
    # AddMedian('上市信息财务信息运营能力指标.xlsx', [x for x in range(3, 7)])
    # AddMedian('上市信息财务信息盈利能力指标.xlsx', [x for x in range(9, 15)])
    # AddMedian('工商基本信息表.xlsx', [x for x in range(1, 2)])
    #
    # print("OneHot")
    # OneHot('工商基本信息表.xlsx', [5, 7, 11])
    # OneHot('海关进出口信用.xlsx', [5])
    #
    # print("WriteCount")
    # WriteExist('一般纳税人.xlsx')
    # WriteCount('债券信息.xlsx')
    # WriteCount('作品著作权.xlsx')
    # WriteCount('资质认证.xlsx')
    # WriteCount('招投标.xlsx')
    # WriteCount('软著著作权.xlsx')
    # WriteCount('融资信息.xlsx')
    # WriteCount('年报-对外投资信息.xlsx')
    # WriteCount('招投标.xlsx')  # 拆成中标、招标
    # WriteCount('招标.xlsx')
    # WriteCount('中标.xlsx')
    # WriteCount('竞品.xlsx')
    # WriteCount('海关进出口信用.xlsx')
    # WriteCount('购地-市场交易-土地转让.xlsx')
    # WriteCount('产品.xlsx')
    # WriteCount('专利.xlsx', islog=True)
    # WriteCount('专利.xlsx', 2, '2016-01-01', islog=True)
    # WriteCount('资质认证.xlsx', 2, '2016-01-01')
    # WriteCount('作品著作权.xlsx', 2, '2016-01-01')
    # WriteCount('债券信息.xlsx')
    # WriteCount('债券信息.xlsx', 4, '2014-01-01')
    #
    # WriteCount('融资信息.xlsx', 2, '2016-01-01')
    # WriteCount('招投标.xlsx', 4, '2016-01-01')
    # WriteCount('招标.xlsx', 4, '2016-01-01')
    # WriteCount('中标.xlsx', 4, '2016-01-01')
    # WriteCount('购地-市场交易-土地转让.xlsx', 1, '2016-01-01')
    # WriteCount('纳税A级年份.xlsx', 1, '2016')
    # WriteCount('年报-股东股权转让.xlsx', 4, '2016')
    # WriteCount('年报-网站或网点信息.xlsx', 1, '2016')
    # WriteCount('年报-对外投资信息.xlsx', 3, '2016')
    # WriteCount('年报-的对外提供保证担保信息.xlsx', 7, '2016')
    # WriteCount('年报-股东股权转让.xlsx')
    # WriteCount('年报-网站或网点信息.xlsx')
    # WriteCount('年报-对外投资信息.xlsx')
    # WriteCount('年报-的对外提供保证担保信息.xlsx')
    #
    # print("AddAverageYear")
    # AddAverageYear('上市公司财务信息-每股指标.xlsx', [x for x in range(3, 10)])
    # AddAverageYear('上市信息财务信息-财务风险指标.xlsx', [x for x in range(3, 7)])
    # AddAverageYear('上市信息财务信息-成长能力指标.xlsx', [x for x in range(17, 23)])
    # AddAverageYear('上市信息财务信息-现金流量表.xlsx', [x for x in range(3, 27)])
    # AddAverageYear('上市信息财务信息资产负债表.xlsx', [x for x in range(3, 23)])

    xlsx_to_csv_pd()
